import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl import load_workbook
from tkinter import *  
from tkinter import messagebox  
 
messagebox.showinfo("Informação", "O programa será executada. Aguarde!") 

# Carrega os dados em um DataFrame do pandas
df = pd.read_excel('Base.xlsx')

# Cria um dicionário de DataFrames para cada equipe
equipes = {}
for equipe in df['Centro de Custo'].unique():
    equipes[equipe] = df[df['Centro de Custo'] == equipe].reset_index(drop=True)

# Cria um arquivo Excel com várias abas, uma para cada equipe
wb = Workbook()
for equipe, df_equipe in equipes.items():
    ws = wb.create_sheet(title=equipe)
    for r in dataframe_to_rows(df_equipe, index=False, header=True):
        ws.append(r)

# Salva o arquivo Excel com as várias abas
wb.save('dados_por_centro_de_custo.xlsx')

# Carrega o arquivo Excel com as abas
wb = load_workbook(filename='dados_por_centro_de_custo.xlsx')

# Percorre as abas do arquivo Excel e salva cada uma em um arquivo separado
for sheet in wb:
    # Cria um novo arquivo Excel
    new_wb = Workbook()
    # Copia a aba atual para o novo arquivo Excel
    new_wb.active = new_wb.create_sheet(sheet.title)
    for row in sheet.rows:
        new_wb.active.append([cell.value for cell in row])
    new_wb.remove(new_wb['Sheet'])
    # Salva o novo arquivo Excel com o nome da aba atual
    new_wb.save(f'{sheet.title}.xlsx')


messagebox.showinfo("Informação", "O programa foi executado!") 